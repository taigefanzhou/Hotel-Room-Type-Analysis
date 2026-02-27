#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Hotel CAD Analyzer - Simple Version for Windows
"""
import sys
import os
from pathlib import Path
from datetime import datetime

def main():
    try:
        from tkinter import Tk, messagebox, filedialog
        import ezdxf
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError as e:
        print(f"Error: {e}")
        input("Press Enter to exit...")
        return 1
    
    root = Tk()
    root.withdraw()
    
    # Select file
    file_path = filedialog.askopenfilename(
        title="Select DXF File",
        filetypes=[("DXF files", "*.dxf"), ("All files", "*.*")]
    )
    
    if not file_path:
        return 0
    
    # Analyze
    try:
        doc = ezdxf.readfile(file_path)
        msp = doc.modelspace()
        
        # Find rooms
        rooms = []
        for entity in msp:
            if entity.dxftype() in ['TEXT', 'MTEXT']:
                try:
                    text = entity.dxf.text if entity.dxftype() == 'TEXT' else entity.text
                    text = text.strip()
                    
                    if '大床' in text:
                        rooms.append({'type': 'King', 'text': text})
                    elif '双床' in text or '标间' in text:
                        rooms.append({'type': 'Double', 'text': text})
                    elif '套房' in text:
                        rooms.append({'type': 'Suite', 'text': text})
                except:
                    pass
        
        if not rooms:
            messagebox.showerror("Error", "No rooms found!")
            return 1
        
        # Output to Desktop
        desktop = Path.home() / "Desktop"
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = desktop / f"CAD_Report_{timestamp}"
        output_dir.mkdir(exist_ok=True)
        
        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Rooms"
        
        ws['A1'] = "Hotel CAD Analysis Report"
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['No.', 'Type', 'Label']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)
        
        for i, room in enumerate(rooms, 1):
            ws.cell(row=i+3, column=1, value=i)
            ws.cell(row=i+3, column=2, value=room['type'])
            ws.cell(row=i+3, column=3, value=room['text'])
        
        excel_path = output_dir / "Room_Analysis.xlsx"
        wb.save(excel_path)
        
        # Create text report
        txt_path = output_dir / "Summary.txt"
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write("Hotel CAD Analysis Report\\n")
            f.write("=" * 50 + "\\n\\n")
            f.write(f"Total Rooms: {len(rooms)}\\n\\n")
            
            stats = {}
            for r in rooms:
                t = r['type']
                stats[t] = stats.get(t, 0) + 1
            
            f.write("Room Types:\\n")
            for t, c in sorted(stats.items()):
                f.write(f"  {t}: {c}\\n")
        
        messagebox.showinfo("Success", f"Reports saved to Desktop:\\n{output_dir}")
        
        # Open folder
        if sys.platform == 'win32':
            os.startfile(str(output_dir))
            
    except Exception as e:
        messagebox.showerror("Error", f"Analysis failed:\\n{str(e)}")
        return 1
    
    return 0

if __name__ == "__main__":
    sys.exit(main())
