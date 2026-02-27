#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Hotel CAD Analyzer - 通用版本
支持各种CAD图纸标注方式
"""
import sys
import os
import re
import math
from pathlib import Path
from datetime import datetime

# ==================== 配置区域 ====================
# 可根据不同图纸调整这些配置

# 房型关键词（支持多种写法）
ROOM_KEYWORDS = {
    '大床房': ['大床', '大床房', 'king', 'king bed', '单人间', '单人房', 'dk', 'kingroom'],
    '双床房': ['双床', '双床房', '标间', '标准间', 'twin', 'twin bed', 'double', '两张床', 'sb', 'twinroom'],
    '套房': ['套房', 'suite', '行政', '豪华', '行政房', '豪华房', 'business', 'deluxe'],
    '总统套房': ['总统', '总统套房', 'presidential', 'president'],
    '家庭房': ['家庭', '家庭房', 'family', '亲子', '亲子房', '三人间'],
}

# 必须过滤的关键词（汇总、统计、公共设施）
MUST_FILTER = [
    '总房间数', '合计', '总计', '统计', '汇总',
    '布草间', '仓库', '仓储', '洗消', '消毒',
    '配电', '空调机房', '弱电', '强电',
    '楼梯', '电梯', '走廊', '过道', '前室',
    '管井', '水井', '电井', '风井',
    '尺寸', '变量', '图例', '说明', '备注',
]

# 判断为汇总信息的规则
def is_summary_text(text):
    """智能判断是否为汇总统计文字"""
    if not text:
        return True
    
    # 包含明确汇总关键词
    if any(kw in text for kw in ['总房间数', '合计', '总计', '统计汇总']):
        return True
    
    # 包含多个冒号和间字（如"大床房：12间\P双床房：9间"）
    if text.count('：') >= 2 and text.count('间') >= 2:
        return True
    
    # 包含换行符和多个房间类型
    if ('\P' in text or '\n' in text) and text.count('房') >= 2:
        return True
    
    # 格式：XX间（纯数字+间）
    if re.match(r'^\d+间$', text.strip()):
        return True
    
    return False

# ==================== 核心算法 ====================

def detect_room_type(text):
    """
    智能识别房型
    返回房型名称或None
    """
    if not text:
        return None
    
    text_clean = text.strip()
    text_lower = text_clean.lower()
    
    # 1. 首先过滤汇总统计
    if is_summary_text(text_clean):
        return None
    
    # 2. 过滤公共设施
    for kw in MUST_FILTER:
        if kw in text_clean:
            return None
    
    # 3. 按优先级识别房型（长的关键词优先）
    # 先检查组合词
    for room_type, keywords in sorted(ROOM_KEYWORDS.items(), key=lambda x: -len(x[0])):
        for kw in keywords:
            if kw in text_clean or kw in text_lower:
                return room_type
    
    return None

def analyze_cad_file(filepath):
    """分析CAD文件"""
    try:
        import ezdxf
    except ImportError:
        return None, "缺少ezdxf库，请运行: pip install ezdxf"
    
    try:
        doc = ezdxf.readfile(filepath)
        msp = doc.modelspace()
    except Exception as e:
        return None, f"无法读取文件: {str(e)}"
    
    # 收集所有文字标注
    all_labels = []
    for entity in msp:
        if entity.dxftype() in ['TEXT', 'MTEXT']:
            try:
                text = entity.dxf.text if entity.dxftype() == 'TEXT' else entity.text
                if text:
                    pos = entity.dxf.insert if hasattr(entity.dxf, 'insert') else (0, 0, 0)
                    all_labels.append({
                        'text': text.strip(),
                        'x': pos[0],
                        'y': pos[1],
                        'raw_text': text.strip()  # 保留原始文本用于调试
                    })
            except:
                pass
    
    # 去重（按位置，允许30单位误差，更精确）
    seen = set()
    unique_labels = []
    for item in all_labels:
        # 使用30单位精度去重
        x_key = round(item['x'] / 30) * 30
        y_key = round(item['y'] / 30) * 30
        key = f"{item['text']}_{x_key}_{y_key}"
        
        if key not in seen and item['text']:
            seen.add(key)
            unique_labels.append(item)
    
    # 识别房间
    rooms = []
    for item in unique_labels:
        text = item['text']
        
        # 使用智能识别
        room_type = detect_room_type(text)
        
        if room_type:
            rooms.append({
                'type': room_type,
                'text': text,
                'x': item['x'],
                'y': item['y'],
                'door_direction': '-',
                'card_position': '-'
            })
    
    return rooms, None

def generate_excel(rooms, output_path):
    """生成Excel报表"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return "缺少openpyxl库，请运行: pip install openpyxl"
    
    wb = Workbook()
    
    # 汇总表
    ws1 = wb.active
    ws1.title = '房型汇总'
    
    # 标题
    ws1['A1'] = '酒店房型统计报表'
    ws1['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws1['A1'].fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    ws1.merge_cells('A1:D1')
    ws1.row_dimensions[1].height = 30
    
    # 生成日期
    ws1['A2'] = f'生成日期: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws1.merge_cells('A2:D2')
    
    # 按楼层统计（按Y坐标分组）
    floors = {}
    for room in rooms:
        y_key = round(room['y'], -2)  # 按100单位分组
        if y_key not in floors:
            floors[y_key] = {}
        t = room['type']
        floors[y_key][t] = floors[y_key].get(t, 0) + 1
    
    row = 4
    total_by_floor = {}
    
    for y, stats in sorted(floors.items(), reverse=True):
        # 楼层标题
        ws1.cell(row=row, column=1, value=f'楼层 (Y≈{y})')
        ws1.cell(row=row, column=1).font = Font(bold=True, color='667eea')
        ws1.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # 表头
        headers = ['房型', '数量', '占比']
        for col, h in enumerate(headers, 2):
            cell = ws1.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E8EAFF', end_color='E8EAFF', fill_type='solid')
        row += 1
        
        # 该层数据
        floor_total = sum(stats.values())
        for t, c in sorted(stats.items()):
            ws1.cell(row=row, column=2, value=t)
            ws1.cell(row=row, column=3, value=c)
            ws1.cell(row=row, column=4, value=f'{c/floor_total*100:.1f}%')
            
            total_by_floor[t] = total_by_floor.get(t, 0) + c
            row += 1
        
        row += 1  # 空行
    
    # 总计
    ws1.cell(row=row, column=1, value='总计')
    ws1.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws1.merge_cells(f'A{row}:D{row}')
    row += 1
    
    grand_total = len(rooms)
    for t, c in sorted(total_by_floor.items()):
        ws1.cell(row=row, column=2, value=t)
        ws1.cell(row=row, column=3, value=c)
        ws1.cell(row=row, column=3).font = Font(bold=True)
        ws1.cell(row=row, column=4, value=f'{c/grand_total*100:.1f}%')
        row += 1
    
    # 合计行
    ws1.cell(row=row, column=2, value='合计')
    ws1.cell(row=row, column=2).font = Font(bold=True, size=12)
    ws1.cell(row=row, column=3, value=grand_total)
    ws1.cell(row=row, column=3).font = Font(bold=True, size=12)
    
    # 设置列宽
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 12
    
    # 明细表
    ws2 = wb.create_sheet('房间明细')
    headers = ['序号', '房型', '原始标注', '坐标位置']
    
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E8EAFF', end_color='E8EAFF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    for i, room in enumerate(rooms, 1):
        ws2.cell(row=i+1, column=1, value=i)
        ws2.cell(row=i+1, column=2, value=room['type'])
        ws2.cell(row=i+1, column=3, value=room['text'])
        ws2.cell(row=i+1, column=4, value=f"X:{room['x']:.0f}, Y:{room['y']:.0f}")
    
    # 设置列宽
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 20
    
    wb.save(output_path)
    return None

def main():
    from tkinter import Tk, messagebox, filedialog
    
    root = Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title='选择CAD文件 (.dxf)',
        filetypes=[('DXF文件', '*.dxf'), ('所有文件', '*.*')]
    )
    
    if not file_path:
        return 0
    
    rooms, error = analyze_cad_file(file_path)
    
    if error:
        messagebox.showerror('错误', error)
        return 1
    
    if not rooms:
        messagebox.showerror('错误', '未识别到房间，请检查CAD文件')
        return 1
    
    # 输出到桌面
    desktop = Path.home() / 'Desktop'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = desktop / f'CAD分析报告_{timestamp}'
    output_dir.mkdir(exist_ok=True)
    
    # 生成Excel
    excel_path = output_dir / '房型分析.xlsx'
    error = generate_excel(rooms, excel_path)
    if error:
        messagebox.showerror('错误', error)
        return 1
    
    # 统计
    stats = {}
    for r in rooms:
        t = r['type']
        stats[t] = stats.get(t, 0) + 1
    
    msg = f'分析完成！\n\n共识别 {len(rooms)} 个房间\n\n'
    for t, c in sorted(stats.items()):
        msg += f'{t}: {c} 间\n'
    msg += f'\n报告已保存到桌面:\n{output_dir}'
    
    messagebox.showinfo('成功', msg)
    
    if sys.platform == 'win32':
        os.startfile(str(output_dir))
    
    return 0

if __name__ == '__main__':
    sys.exit(main())
