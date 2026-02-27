#!/usr/bin/env python3
"""
酒店 CAD 平面图分析器
自动识别房型、统计房间数量、分析开门方向、生成 Excel 报表
"""

import os
import sys
import re
import math
import argparse
from pathlib import Path
from datetime import datetime

try:
    import ezdxf
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import pandas as pd
except ImportError as e:
    print(f"❌ 缺少依赖: {e}")
    print("请运行: pip3 install ezdxf openpyxl pandas pillow")
    sys.exit(1)


# 房型关键词映射
ROOM_TYPE_PATTERNS = {
    "大床房": [r"大床", r"单人间", r"king", r"single", r"大床房"],
    "双床房": [r"双床", r"标间", r"标准间", r"twin", r"double", r"双人间"],
    "套房": [r"套房", r"suite", r"行政", r"行政房", r"豪华套房"],
    "总统套房": [r"总统", r"presidential", r"president"],
    "家庭房": [r"家庭", r"family", r"亲子"],
    "无障碍房": [r"无障碍", r"残疾人", r"accessible", r"wheelchair"],
}


def detect_room_type(text):
    """根据文本内容识别房型"""
    text_lower = text.lower()
    
    # 按优先级检查
    for room_type, patterns in ROOM_TYPE_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, text_lower):
                return room_type
    
    return "其他"


def extract_room_number(text):
    """提取房间号（如 801, 1002, 1-DT6 等）"""
    # 匹配 3-4 位纯数字（标准房间号）
    matches = re.findall(r'\b(\d{3,4})\b', text)
    if matches:
        return matches[0]
    
    # 匹配类似 1-DT6, 2-LT3 的编号格式
    matches = re.findall(r'\b(\d+[\-\.]?[A-Z]{1,2}\d*)\b', text)
    if matches:
        return matches[0]
    
    # 匹配单个字母编号（A, B, C...）
    matches = re.findall(r'\b([A-Z])\b', text)
    if matches and len(text) <= 3:
        return matches[0]
    
    return None


def calculate_door_direction(start_angle, end_angle):
    """计算门开启方向"""
    # 计算弧线中点角度
    mid_angle = (start_angle + end_angle) / 2
    if end_angle < start_angle:
        mid_angle = (start_angle + end_angle + 360) / 2
        if mid_angle > 360:
            mid_angle -= 360
    
    # 判断主要方向
    if 45 <= mid_angle < 135:
        return "向上开启", "内侧"
    elif 135 <= mid_angle < 225:
        return "向左开启", "左侧"
    elif 225 <= mid_angle < 315:
        return "向下开启", "内侧"
    else:  # 315-360 或 0-45
        return "向右开启", "右侧"


def calculate_distance(pos1, pos2):
    """计算两点距离"""
    return math.sqrt((pos1[0] - pos2[0])**2 + (pos1[1] - pos2[1])**2)


def analyze_doors(msp):
    """分析门信息"""
    doors = []
    
    for entity in msp:
        if entity.dxftype() == 'ARC':
            try:
                # 从 FF-门 图层找门弧线
                if entity.dxf.layer == 'FF-门':
                    center = entity.dxf.center
                    radius = entity.dxf.radius
                    start_angle = entity.dxf.start_angle
                    end_angle = entity.dxf.end_angle
                    
                    # 过滤掉太小的装饰弧线
                    if radius < 15:
                        continue
                    
                    # 计算开门方向
                    direction, card_position = calculate_door_direction(start_angle, end_angle)
                    
                    doors.append({
                        'center': (center.x, center.y),
                        'radius': radius,
                        'start_angle': start_angle,
                        'end_angle': end_angle,
                        'direction': direction,
                        'card_position': card_position,  # 插卡取电建议位置
                    })
            except Exception:
                pass
    
    return doors


def match_doors_to_rooms(rooms, doors):
    """将门与最近的房间关联"""
    for room in rooms:
        room_pos = room['position']
        closest_door = None
        min_distance = float('inf')
        
        for door in doors:
            dist = calculate_distance(room_pos, door['center'])
            if dist < min_distance:
                min_distance = dist
                closest_door = door
        
        if closest_door and min_distance < 200:  # 距离阈值 200 单位
            room['door_direction'] = closest_door['direction']
            room['card_position'] = closest_door['card_position']
            room['door_distance'] = round(min_distance, 1)
        else:
            room['door_direction'] = "未知"
            room['card_position'] = "未知"
            room['door_distance'] = "-"
    
    return rooms


def analyze_cad_file(filepath):
    """分析 CAD 文件，提取房间信息和门方向"""
    print(f"📁 正在分析: {filepath}")
    
    if not os.path.exists(filepath):
        print(f"❌ 文件不存在: {filepath}")
        return None
    
    try:
        doc = ezdxf.readfile(filepath)
    except Exception as e:
        print(f"⚠️  无法直接读取: {e}")
        print("提示: 如果是 .dwg 文件，请先转换为 .dxf 格式")
        return None
    
    msp = doc.modelspace()
    
    # ========== 第一步：分析房间标注 ==========
    print("🔍 正在扫描房间标注...")
    
    text_entities = []
    for entity in msp:
        if entity.dxftype() in ['TEXT', 'MTEXT']:
            try:
                text = entity.dxf.text if entity.dxftype() == 'TEXT' else entity.text
                if text:
                    text_entities.append({
                        'text': text.strip(),
                        'type': entity.dxftype(),
                        'position': (entity.dxf.insert[0], entity.dxf.insert[1]) if hasattr(entity.dxf, 'insert') else (0, 0)
                    })
            except Exception:
                pass
    
    print(f"📊 找到 {len(text_entities)} 个文字标注")
    
    # 去重
    seen = set()
    unique_entities = []
    for item in text_entities:
        pos = item['position']
        key = f"{item['text']}_{round(pos[0], -1)}_{round(pos[1], -1)}"
        if key not in seen and item['text']:
            seen.add(key)
            unique_entities.append(item)
    
    # 识别房间
    rooms = []
    for item in unique_entities:
        text = item['text']
        room_type = detect_room_type(text)
        room_number = extract_room_number(text)
        
        if room_type != "其他":
            rooms.append({
                'raw_text': text,
                'room_type': room_type,
                'room_number': room_number,
                'position': item['position'],
                'door_direction': "未知",
                'card_position': "未知",
                'door_distance': "-"
            })
    
    print(f"✅ 识别到 {len(rooms)} 个房间")
    
    # ========== 第二步：分析门方向 ==========
    print("🔍 正在分析门开启方向...")
    doors = analyze_doors(msp)
    print(f"✅ 找到 {len(doors)} 个门")
    
    # ========== 第三步：关联门和房间 ==========
    if doors:
        print("🔗 正在关联门与房间...")
        rooms = match_doors_to_rooms(rooms, doors)
    
    return rooms


def generate_excel_report(rooms, output_path):
    """生成 Excel 报表"""
    print(f"📑 正在生成 Excel 报表: {output_path}")
    
    # 统计房型
    room_stats = {}
    for room in rooms:
        room_type = room['room_type']
        room_stats[room_type] = room_stats.get(room_type, 0) + 1
    
    wb = Workbook()
    
    # ========== Sheet 1: 汇总统计 ==========
    ws_summary = wb.active
    ws_summary.title = "房型汇总"
    
    # 标题
    ws_summary['A1'] = "酒店房型统计报表"
    ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.merge_cells('A1:C1')
    ws_summary.row_dimensions[1].height = 30
    
    # 生成日期
    ws_summary['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A2'].font = Font(italic=True, color="666666")
    ws_summary.merge_cells('A2:C2')
    
    # 表头
    headers = ['房型', '数量', '占比']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 数据
    total_rooms = sum(room_stats.values())
    row = 5
    for room_type in sorted(room_stats.keys()):
        count = room_stats[room_type]
        percentage = f"{count/total_rooms*100:.1f}%" if total_rooms > 0 else "0%"
        
        ws_summary.cell(row=row, column=1, value=room_type).alignment = Alignment(horizontal='center')
        ws_summary.cell(row=row, column=2, value=count).alignment = Alignment(horizontal='center')
        ws_summary.cell(row=row, column=3, value=percentage).alignment = Alignment(horizontal='center')
        row += 1
    
    # 合计行
    ws_summary.cell(row=row, column=1, value="合计").font = Font(bold=True)
    ws_summary.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=2, value=total_rooms).font = Font(bold=True)
    ws_summary.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=3, value="100%").font = Font(bold=True)
    ws_summary.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    
    # 设置列宽
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    
    # ========== Sheet 2: 房间明细（含门方向） ==========
    ws_detail = wb.create_sheet("房间明细")
    
    # 标题
    ws_detail['A1'] = "房间明细列表（含开门方向）"
    ws_detail['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_detail['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws_detail['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_detail.merge_cells('A1:G1')
    ws_detail.row_dimensions[1].height = 25
    
    # 表头
    detail_headers = ['序号', '房间号', '房型', '开门方向', '插卡取电位置', '门距(m)', '原始标注']
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 数据
    for idx, room in enumerate(rooms, 1):
        room_num = room.get('room_number') or "-"
        door_dist = room.get('door_distance') if room.get('door_distance') != "-" else "-"
        if isinstance(door_dist, (int, float)):
            door_dist = f"{door_dist/100:.2f}"  # 转换为米
        
        ws_detail.cell(row=idx+3, column=1, value=idx).alignment = Alignment(horizontal='center')
        ws_detail.cell(row=idx+3, column=2, value=room_num).alignment = Alignment(horizontal='center')
        ws_detail.cell(row=idx+3, column=3, value=room['room_type']).alignment = Alignment(horizontal='center')
        ws_detail.cell(row=idx+3, column=4, value=room['door_direction']).alignment = Alignment(horizontal='center')
        ws_detail.cell(row=idx+3, column=5, value=room['card_position']).alignment = Alignment(horizontal='center')
        ws_detail.cell(row=idx+3, column=6, value=door_dist).alignment = Alignment(horizontal='center')
        ws_detail.cell(row=idx+3, column=7, value=room['raw_text']).alignment = Alignment(horizontal='left')
    
    # 设置列宽
    ws_detail.column_dimensions['A'].width = 8
    ws_detail.column_dimensions['B'].width = 12
    ws_detail.column_dimensions['C'].width = 12
    ws_detail.column_dimensions['D'].width = 14
    ws_detail.column_dimensions['E'].width = 16
    ws_detail.column_dimensions['F'].width = 12
    ws_detail.column_dimensions['G'].width = 20
    
    # 保存文件
    wb.save(output_path)
    print(f"✅ Excel 报表已生成: {output_path}")
    
    return room_stats


def generate_summary_text(rooms, output_path):
    """生成文本摘要"""
    room_stats = {}
    for room in rooms:
        room_type = room['room_type']
        room_stats[room_type] = room_stats.get(room_type, 0) + 1
    
    total = sum(room_stats.values())
    
    lines = [
        "=" * 70,
        "              酒店房型分析报表（含开门方向）",
        "=" * 70,
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "【房型统计】",
        "-" * 50,
    ]
    
    for room_type in sorted(room_stats.keys()):
        count = room_stats[room_type]
        percentage = count/total*100 if total > 0 else 0
        lines.append(f"  {room_type}: {count} 间 ({percentage:.1f}%)")
    
    lines.extend([
        "-" * 50,
        f"  总房间数: {total} 间",
        "",
        "【房间明细】",
        "-" * 70,
        f"{'序号':<6}{'房号':<10}{'房型':<10}{'开门方向':<12}{'插卡位置':<12}{'距离':<8}原始标注",
        "-" * 70,
    ])
    
    for idx, room in enumerate(rooms, 1):
        room_num = room.get('room_number') or "-"
        door_dist = room.get('door_distance') if room.get('door_distance') != "-" else "-"
        if isinstance(door_dist, (int, float)):
            door_dist = f"{door_dist/100:.2f}m"
        
        lines.append(
            f"{idx:<6}{room_num:<10}{room['room_type']:<10}{room['door_direction']:<12}"
            f"{room['card_position']:<12}{door_dist:<8}{room['raw_text']}"
        )
    
    lines.append("=" * 70)
    
    # 插卡取电统计
    lines.extend([
        "",
        "【插卡取电位置统计】",
        "-" * 50,
    ])
    
    position_stats = {}
    for room in rooms:
        pos = room['card_position']
        position_stats[pos] = position_stats.get(pos, 0) + 1
    
    for pos, count in sorted(position_stats.items()):
        lines.append(f"  {pos}: {count} 间")
    
    lines.append("=" * 70)
    
    # 保存
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    
    print(f"✅ 文本摘要已生成: {output_path}")
    print("\n" + '\n'.join(lines))


def main():
    parser = argparse.ArgumentParser(description='酒店 CAD 平面图分析器（含开门方向）')
    parser.add_argument('--input', '-i', required=True, help='CAD 文件路径 (.dxf)')
    parser.add_argument('--output', '-o', default='./output', help='输出目录')
    args = parser.parse_args()
    
    input_path = os.path.expanduser(args.input)
    output_dir = os.path.expanduser(args.output)
    
    os.makedirs(output_dir, exist_ok=True)
    
    # 分析 CAD 文件
    rooms = analyze_cad_file(input_path)
    
    if not rooms:
        print("❌ 未能识别到任何房间信息")
        return 1
    
    # 生成文件名
    base_name = Path(input_path).stem
    excel_path = os.path.join(output_dir, f"{base_name}_房型分析.xlsx")
    txt_path = os.path.join(output_dir, f"{base_name}_分析摘要.txt")
    
    # 生成报表
    generate_excel_report(rooms, excel_path)
    generate_summary_text(rooms, txt_path)
    
    print("\n" + "=" * 50)
    print("✅ 分析完成！")
    print(f"📊 Excel 报表: {excel_path}")
    print(f"📄 文本摘要: {txt_path}")
    print("=" * 50)
    
    return 0


if __name__ == '__main__':
    sys.exit(main())