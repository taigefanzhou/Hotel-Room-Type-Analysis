#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Hotel CAD Analyzer - 酒店房型分析器
智能识别房型，过滤汇总信息，生成Excel报表
"""
import sys
import os
import re
import math
from pathlib import Path
from datetime import datetime

def analyze_cad_file(filepath):
    """分析CAD文件，识别房间信息"""
    try:
        import ezdxf
    except ImportError:
        return None, "缺少ezdxf库，请运行: pip install ezdxf"
    
    try:
        doc = ezdxf.readfile(filepath)
        msp = doc.modelspace()
    except Exception as e:
        return None, f"无法读取文件: {str(e)}"
    
    # 收集所有文字标注
    all_labels = []
    for entity in msp:
        if entity.dxftype() in ['TEXT', 'MTEXT']:
            try:
                text = entity.dxf.text if entity.dxftype() == 'TEXT' else entity.text
                if text:
                    pos = entity.dxf.insert if hasattr(entity.dxf, 'insert') else (0, 0, 0)
                    all_labels.append({
                        'text': text.strip(),
                        'x': pos[0],
                        'y': pos[1]
                    })
            except:
                pass
    
    # 去重（按位置，50单位误差）
    seen = set()
    unique_labels = []
    for item in all_labels:
        x_key = round(item['x'] / 50) * 50
        y_key = round(item['y'] / 50) * 50
        key = f"{item['text']}_{x_key}_{y_key}"
        if key not in seen and item['text']:
            seen.add(key)
            unique_labels.append(item)
    
    # 识别房间
    rooms = []
    for item in unique_labels:
        text = item['text']
        
        # 过滤汇总统计
        if '总房间数' in text or '合计' in text or text.count('间') >= 3:
            continue
        
        # 过滤非房间
        non_room_words = ['布草间', '仓库', '仓储', '洗消', '消毒', '配电', 
                         '空调', '弱电', '强电', '楼梯', '电梯', '走廊', 
                         '过道', '前室', '管井', '水井', '电井', '风井',
                         '尺寸', '变量', '图例', '说明']
        if any(w in text for w in non_room_words):
            continue
        
        # 识别房型
        room_type = None
        
        if '总统' in text:
            room_type = '总统套房'
        elif '套房' in text or '行政' in text or '豪华' in text:
            room_type = '套房'
        elif '家庭' in text or '亲子' in text:
            room_type = '家庭房'
        elif '双床' in text or '标间' in text:
            room_type = '双床房'
        elif '大床' in text or '单人间' in text:
            room_type = '大床房'
        
        if room_type:
            rooms.append({
                'type': room_type,
                'text': text,
                'x': item['x'],
                'y': item['y'],
                'door_direction': '-',
                'card_position': '-'
            })
    
    # 分析门方向
    doors = []
    for entity in msp:
        if entity.dxftype() == 'ARC':
            try:
                if entity.dxf.layer == 'FF-门':
                    center = entity.dxf.center
                    radius = entity.dxf.radius
                    
                    if radius < 15:
                        continue
                    
                    start = entity.dxf.start_angle
                    end = entity.dxf.end_angle
                    
                    mid = (start + end) / 2
                    if end < start:
                        mid = (start + end + 360) / 2
                        if mid > 360:
                            mid -= 360
                    
                    if 45 <= mid < 135:
                        direction, card_pos = '向上开启', '内侧'
                    elif 135 <= mid < 225:
                        direction, card_pos = '向左开启', '左侧'
                    elif 225 <= mid < 315:
                        direction, card_pos = '向下开启', '内侧'
                    else:
                        direction, card_pos = '向右开启', '右侧'
                    
                    doors.append({
                        'center': (center.x, center.y),
                        'direction': direction,
                        'card_pos': card_pos
                    })
            except:
                pass
    
    # 关联门和房间
    for room in rooms:
        rx, ry = room['x'], room['y']
        closest = None
        min_dist = float('inf')
        
        for door in doors:
            dx, dy = door['center']
            dist = math.sqrt((rx-dx)**2 + (ry-dy)**2)
            if dist < min_dist:
                min_dist = dist
                closest = door
        
        if closest and min_dist < 200:
            room['door_direction'] = closest['direction']
            room['card_position'] = closest['card_pos']
    
    return rooms, None

def generate_excel(rooms, output_path):
    """生成Excel报表"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return "缺少openpyxl库，请运行: pip install openpyxl"
    
    wb = Workbook()
    
    # 汇总表
    ws1 = wb.active
    ws1.title = '房型汇总'
    
    ws1['A1'] = '酒店房型统计报表'
    ws1['A1'].font = Font(size=14, bold=True)
    ws1.merge_cells('A1:C1')
    
    # 按楼层统计
    floors = {}
    for room in rooms:
        y_key = round(room['y'], -2)
        if y_key not in floors:
            floors[y_key] = {}
        t = room['type']
        floors[y_key][t] = floors[y_key].get(t, 0) + 1
    
    row = 3
    for y, stats in sorted(floors.items(), reverse=True):
        ws1.cell(row=row, column=1, value=f'楼层 Y={y}')
        ws1.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        for t, c in sorted(stats.items()):
            ws1.cell(row=row, column=2, value=t)
            ws1.cell(row=row, column=3, value=c)
            row += 1
        row += 1
    
    # 总计
    ws1.cell(row=row, column=1, value='总计')
    ws1.cell(row=row, column=1).font = Font(bold=True)
    
    total_stats = {}
    for r in rooms:
        t = r['type']
        total_stats[t] = total_stats.get(t, 0) + 1
    
    for t, c in sorted(total_stats.items()):
        row += 1
        ws1.cell(row=row, column=2, value=t)
        ws1.cell(row=row, column=3, value=c)
    
    row += 1
    ws1.cell(row=row, column=2, value='合计')
    ws1.cell(row=row, column=2).font = Font(bold=True)
    ws1.cell(row=row, column=3, value=len(rooms))
    ws1.cell(row=row, column=3).font = Font(bold=True)
    
    # 明细表
    ws2 = wb.create_sheet('房间明细')
    headers = ['序号', '房型', '开门方向', '插卡位置', '坐标']
    
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for i, room in enumerate(rooms, 1):
        ws2.cell(row=i+1, column=1, value=i)
        ws2.cell(row=i+1, column=2, value=room['type'])
        ws2.cell(row=i+1, column=3, value=room['door_direction'])
        ws2.cell(row=i+1, column=4, value=room['card_position'])
        ws2.cell(row=i+1, column=5, value=f"({room['x']:.0f}, {room['y']:.0f})")
    
    wb.save(output_path)
    return None

def main():
    from tkinter import Tk, messagebox, filedialog
    
    root = Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title='选择CAD文件 (.dxf)',
        filetypes=[('DXF文件', '*.dxf'), ('所有文件', '*.*')]
    )
    
    if not file_path:
        return 0
    
    rooms, error = analyze_cad_file(file_path)
    
    if error:
        messagebox.showerror('错误', error)
        return 1
    
    if not rooms:
        messagebox.showerror('错误', '未识别到房间')
        return 1
    
    # 输出到桌面
    desktop = Path.home() / 'Desktop'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = desktop / f'CAD分析报告_{timestamp}'
    output_dir.mkdir(exist_ok=True)
    
    # 生成Excel
    excel_path = output_dir / '房型分析.xlsx'
    error = generate_excel(rooms, excel_path)
    if error:
        messagebox.showerror('错误', error)
        return 1
    
    # 统计
    stats = {}
    for r in rooms:
        t = r['type']
        stats[t] = stats.get(t, 0) + 1
    
    msg = f'分析完成！\n\n共识别 {len(rooms)} 个房间\n\n'
    for t, c in sorted(stats.items()):
        msg += f'{t}: {c} 间\n'
    msg += f'\n报告已保存到桌面:\n{output_dir}'
    
    messagebox.showinfo('成功', msg)
    
    if sys.platform == 'win32':
        os.startfile(str(output_dir))
    
    return 0

if __name__ == '__main__':
    sys.exit(main())
